from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import datetime

def get_google_results(keyword):
    driver = webdriver.Chrome(executable_path="C:\Users\YasirHossain\Downloads\chromedriver\chromedriver.exe")  
    driver.get("https://www.google.com")

    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(keyword)
    search_box.submit()

    try:
        wait = WebDriverWait(driver, 10)
        results = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#search .g")))
    except:
        results = []

    options = [result.text for result in results]
    driver.quit()

    return options

def process_data(keyword, options, sheet):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Keyword"
    sheet["B1"] = "Longest Option"
    sheet["C1"] = "Shortest Option"
    
    longest_option = max(options, key=len)  
    shortest_option = min(options, key=len)  
    sheet["A2"] = keyword
    sheet["B2"] = longest_option
    sheet["C2"] = shortest_option

    workbook.save("search_results.xlsx")

def main():
    workbook = openpyxl.load_workbook("keywords.xlsx")  
    sheet = workbook.active

    today = datetime.date.today().weekday()  
    
    for row in range(2, sheet.max_row + 1):
        keyword = sheet.cell(row, 1).value
        if today == 0 and sheet.cell(row, 2).value == "Monday" or \
           today == 1 and sheet.cell(row, 3).value == "Tuesday" or \
           today == 2 and sheet.cell(row, 4).value == "Wednesday" or \
           today == 3 and sheet.cell(row, 5).value == "Thursday" or \
           today == 4 and sheet.cell(row, 6).value == "Friday" or \
           today == 5 and sheet.cell(row, 7).value == "Saturday" or \
           today == 6 and sheet.cell(row, 8).value == "Sunday":
            options = get_google_results(keyword)
            process_data(keyword, options, sheet)

    workbook.save("results.xlsx")

if __name__ == "__main__":
    main()
